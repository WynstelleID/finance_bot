# excel_generator.py
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment # Add this line
import openpyxl.styles # Add this line to make openpyxl.styles available directly
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

# Assuming TransactionType is available (e.g., passed or imported)
# For a standalone script, you might need to import it if not passed,
# but for this structure, it's passed via the app.py call.
# Let's add a placeholder for it, assuming it's available in the context of the call
from models import TransactionType # Import TransactionType

def generate_excel_report(transactions, start_date=None, end_date=None, user_id=""):
    """
    Generates an Excel report from a list of transaction objects.

    Args:
        transactions (list): A list of Transaction objects.
        start_date (datetime): The start date for the report period (optional).
        end_date (datetime): The end date for the report period (optional).
        user_id (str): The WhatsApp number for whom the report is generated.

    Returns:
        io.BytesIO: A BytesIO object containing the Excel file data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # Define headers
    headers = ["Date", "Type", "Category", "Amount", "Notes"]
    ws.append(headers)

    # Apply header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = "000080" # Dark blue
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = openpyxl.styles.PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
        cell.border = header_border
        cell.alignment = header_alignment
        # Adjust column width based on content
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Add data rows
    for transaction in transactions:
        date_str = transaction.transaction_date.strftime('%Y-%m-%d %H:%M:%S')
        category_name = transaction.category.name if transaction.category else "N/A"
        row_data = [
            date_str,
            transaction.type.value.capitalize(),
            category_name,
            transaction.amount,
            transaction.notes if transaction.notes else ""
        ]
        ws.append(row_data)

    # Apply cell styling for data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(vertical='center')

            # Highlight income/expense
            if ws.cell(row=row_idx, column=2).value == "Income":
                cell.font = Font(color="006400") # Dark green
            elif ws.cell(row=row_idx, column=2).value == "Expense":
                cell.font = Font(color="FF0000") # Red
            elif ws.cell(row=row_idx, column=2).value == "Asset_adjustment":
                cell.font = Font(color="0000FF") # Blue


    # Add a summary section
    ws.append([]) # Empty row for spacing
    ws.append([]) # Empty row for spacing

    # Calculate totals
    total_income = sum(t.amount for t in transactions if t.type == TransactionType.INCOME)
    total_expense = sum(t.amount for t in transactions if t.type == TransactionType.EXPENSE)
    total_asset_adjustment = sum(t.amount for t in transactions if t.type == TransactionType.ASSET_ADJUSTMENT)
    net_flow = total_income - total_expense + total_asset_adjustment

    summary_data = [
        ["Report Period:", f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}" if start_date and end_date else "All Time"],
        ["Total Income:", total_income],
        ["Total Expenses:", total_expense],
        ["Total Asset Adjustments:", total_asset_adjustment],
        ["Net Flow (Income - Expenses + Adjustments):", net_flow]
    ]

    for row in summary_data:
        ws.append(row)

    # Style summary section
    for row_idx in range(ws.max_row - len(summary_data) + 1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Save to a BytesIO object
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0) # Rewind to the beginning of the stream
    return excel_buffer
